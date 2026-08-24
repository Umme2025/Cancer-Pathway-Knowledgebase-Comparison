import pandas as pd
from pathlib import Path
files = [Path('C:\\Users\\aps211\\OneDrive - The University of Texas-Rio Grande Valley\\Vidio Presentation\\Research_Own\\Pathway Comparison\\data_processed\\all_pathway_resources_master.csv'), Path('C:\\Users\\aps211\\OneDrive - The University of Texas-Rio Grande Valley\\Vidio Presentation\\Research_Own\\Pathway Comparison\\data_processed\\BRCA_PRISM_all_pathway_resources_master.csv')]
for file_path in files:
    print('\n' + '=' * 80)
    print('FILE:', file_path.name)
    df = pd.read_csv(file_path)
    df['Gene_Symbol_clean'] = df['Gene_Symbol'].astype(str).str.strip().str.upper()
    print('Shape:', df.shape)
    print('\nUnique genes by database:')
    print(df.groupby('Database')['Gene_Symbol_clean'].nunique().sort_index())
    reactome_count = df.loc[df['Database'].astype(str).str.strip().str.lower() == 'reactome', 'Gene_Symbol_clean'].nunique()
    print('\n>>> EXACT REACTOME UNIQUE GENES =', reactome_count)
from pathlib import Path
base = Path('C:\\Users\\aps211\\OneDrive - The University of Texas-Rio Grande Valley')
matches = list(base.rglob('*pathway*resources*master*.csv'))
print('Found files:')
for f in matches:
    print(f)
from pathlib import Path
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
PROJECT_ROOT = Path(__file__).resolve().parents[1]
MASTER_FILE = PROJECT_ROOT / 'data_processed' / 'BRCA_PRISM_all_pathway_resources_master.csv'
FIGURES_DIR = PROJECT_ROOT / 'figures'
RESULTS_DIR = PROJECT_ROOT / 'results'
FIGURES_DIR.mkdir(parents=True, exist_ok=True)
RESULTS_DIR.mkdir(parents=True, exist_ok=True)
master = pd.read_csv(MASTER_FILE)
print('Master shape:', master.shape)
print('Columns:', master.columns.tolist())
resource_order = ['LCPathways', 'KEGG', 'Reactome', 'WikiPathways', 'MSigDB Hallmark']
if 'Gene_Symbol_clean' in master.columns:
    gene_col = 'Gene_Symbol_clean'
else:
    gene_col = 'Gene_Symbol'
    master[gene_col] = master[gene_col].astype(str).str.strip().str.upper()
print('Using gene column:', gene_col)
master = master.dropna(subset=[gene_col, 'Database']).copy()
exact_counts = master.groupby('Database')[gene_col].nunique().reindex(resource_order)
print('\nExact unique genes by database:')
print(exact_counts)
reactome_count = exact_counts['Reactome']
print(f'\n>>> EXACT REACTOME UNIQUE GENES = {reactome_count}')
assert reactome_count == 11983, f'Reactome count is {reactome_count}, expected 11983'
membership = master[[gene_col, 'Database']].drop_duplicates().assign(Present=1).pivot(index=gene_col, columns='Database', values='Present').fillna(0).astype(int)
membership = membership[resource_order]
membership['Resource_Count'] = membership.sum(axis=1)
print('\nTotal genes in union:', len(membership))
combination_counts = membership.groupby(resource_order, as_index=False).size().rename(columns={'size': 'Gene_Count'})
combination_counts = combination_counts.sort_values('Gene_Count', ascending=False).reset_index(drop=True)

def readable_combination(row):
    present = [r for r in resource_order if row[r] == 1]
    return ' + '.join(present)
combination_counts['Combination'] = combination_counts.apply(readable_combination, axis=1)
top_n = 15
plot_df = combination_counts.head(top_n).copy().reset_index(drop=True)
plot_df['Rank'] = np.arange(1, len(plot_df) + 1)
rank_key = plot_df[['Rank', 'Combination', 'Gene_Count']].copy()
print('\nTop intersection rank key:')
print(rank_key)
rank_key.to_csv(RESULTS_DIR / 'BRCA_PRISM_gene_overlap_rank_key.csv', index=False)
combination_counts.to_csv(RESULTS_DIR / 'BRCA_PRISM_gene_exact_intersection_counts.csv', index=False)
set_sizes = [membership[r].sum() for r in resource_order]
set_size_df = pd.DataFrame({'Database': resource_order, 'Set_Size': set_sizes})
print('\nSet sizes used in figure:')
print(set_size_df)
assert set_size_df.loc[set_size_df['Database'] == 'Reactome', 'Set_Size'].iloc[0] == 11983
resource_colors = {'LCPathways': '#1f77b4', 'KEGG': '#ff7f0e', 'Reactome': '#2ca02c', 'WikiPathways': '#d62728', 'MSigDB Hallmark': '#9467bd'}
inactive_color = '#d9d9d9'
bar_color = '#4c78a8'
line_color = '#4d4d4d'
n = len(plot_df)
x = np.arange(n)
y_positions = np.arange(len(resource_order))[::-1]
fig = plt.figure(figsize=(16, 9))
ax_left = fig.add_axes([0.04, 0.12, 0.13, 0.34])
ax_matrix = fig.add_axes([0.27, 0.12, 0.69, 0.34])
ax_bar = fig.add_axes([0.26, 0.54, 0.69, 0.36])
ax_bar.bar(x, plot_df['Gene_Count'], color=bar_color, edgecolor='black', linewidth=0.6)
ax_bar.set_ylabel('Number of genes', fontsize=12)
ax_bar.set_title('Gene overlap across five pathway resources', fontsize=15, pad=10)
ax_bar.set_xticks(x)
ax_bar.set_xticklabels(plot_df['Rank'].astype(str), fontsize=10)
ax_bar.set_xlabel('Intersection rank', fontsize=12)
for i, count in enumerate(plot_df['Gene_Count']):
    ax_bar.text(i, count + max(plot_df['Gene_Count']) * 0.01, f'{count:,}', ha='center', va='bottom', fontsize=9)
ax_bar.spines['top'].set_visible(False)
ax_bar.spines['right'].set_visible(False)
for col_idx, (_, row) in enumerate(plot_df.iterrows()):
    active_positions = []
    for resource_idx, resource in enumerate(resource_order):
        y = y_positions[resource_idx]
        if row[resource] == 1:
            ax_matrix.scatter(col_idx, y, s=90, color=resource_colors[resource], edgecolor='black', linewidth=0.4, zorder=3)
            active_positions.append(y)
        else:
            ax_matrix.scatter(col_idx, y, s=55, color=inactive_color, edgecolor='none', zorder=2)
    if len(active_positions) > 1:
        ax_matrix.plot([col_idx, col_idx], [min(active_positions), max(active_positions)], color=line_color, linewidth=1.4, zorder=1)
ax_matrix.set_yticks(y_positions)
ax_matrix.set_yticklabels(resource_order, fontsize=12)
ax_matrix.set_xticks(x)
ax_matrix.set_xticklabels(plot_df['Rank'].astype(str), fontsize=11)
ax_matrix.set_xlabel('Intersection rank', fontsize=12)
ax_matrix.set_xlim(-0.8, n - 0.2)
ax_matrix.set_ylim(-0.6, len(resource_order) - 0.4)
ax_matrix.spines['top'].set_visible(False)
ax_matrix.spines['right'].set_visible(False)
ax_matrix.tick_params(axis='y', pad=2)
for y in y_positions:
    ax_matrix.axhline(y=y, color='#f0f0f0', linewidth=0.8, zorder=0)
ax_left.barh(y_positions, set_sizes, color=bar_color, edgecolor='black', linewidth=0.6)
ax_left.set_yticks([])
ax_left.set_xlabel('Number of unique genes', fontsize=11)
ax_left.set_xlim(max(set_sizes) * 1.05, 0)
for y, count in zip(y_positions, set_sizes):
    offset = max(150, int(count * 0.03))
    ax_left.text(count - offset, y, f'{count:,}', va='center', ha='left', fontsize=10)
ax_left.spines['top'].set_visible(False)
ax_left.spines['right'].set_visible(False)
png_file = FIGURES_DIR / 'BRCA_PRISM_gene_overlap_upset_clean_fullnames.png'
pdf_file = FIGURES_DIR / 'BRCA_PRISM_gene_overlap_upset_clean_fullnames.pdf'
plt.savefig(png_file, dpi=300, bbox_inches='tight')
plt.savefig(pdf_file, bbox_inches='tight')
plt.show()
print('\nSaved figures:')
print(png_file)
print(pdf_file)
support_summary = membership['Resource_Count'].value_counts().sort_index().rename_axis('Number_of_Resources').reset_index(name='Number_of_Genes')
support_summary['Percentage'] = support_summary['Number_of_Genes'] / len(membership) * 100
support_summary.to_csv(RESULTS_DIR / 'BRCA_PRISM_gene_resource_support_summary.csv', index=False)
print('\nGene support summary:')
print(support_summary.round(2))
print('\nINTERPRETATION:')
print('The x-axis labels 1, 2, 3, ..., 15 are INTERSECTION RANKS.')
print('Rank 1 = largest exact gene-overlap combination.')
print('Rank 2 = second largest exact gene-overlap combination.')
print('Rank 3 = third largest exact gene-overlap combination.')
print('Use the saved rank key file to identify each combination:')
print(RESULTS_DIR / 'BRCA_PRISM_gene_overlap_rank_key.csv')
from pathlib import Path
PROJECT_ROOT = Path(__file__).resolve().parents[1]
print('PROJECT ROOT:')
print(PROJECT_ROOT)
keywords = ['review', 'jaccard', 'intersection', 'family', 'pam50', 'importance', 'ablation', 'cptac', 'concordance', 'sensitivity']
for keyword in keywords:
    print('\n' + '=' * 80)
    print('KEYWORD:', keyword.upper())
    matches = [f for f in PROJECT_ROOT.rglob('*') if f.is_file() and keyword.lower() in f.name.lower() and (f.suffix.lower() in ['.csv', '.xlsx', '.xls'])]
    for f in matches:
        print(f)
from pathlib import Path
import pandas as pd
import itertools
PROJECT_ROOT = Path(__file__).resolve().parents[1]
RESULTS_DIR = PROJECT_ROOT / 'results'
DATA_DIR = PROJECT_ROOT / 'data_processed'
SUPP_DIR = PROJECT_ROOT / 'supplementary'
SUPP_DIR.mkdir(parents=True, exist_ok=True)
MASTER_FILE = DATA_DIR / 'BRCA_PRISM_all_pathway_resources_master.csv'
S1_file = RESULTS_DIR / 'BRCA_PRISM_step6C_BRCA_family_manual_review_COMPLETED-(S1).csv'
S1 = pd.read_csv(S1_file)
print('S1:', S1.shape)
print(S1.columns.tolist())
master = pd.read_csv(MASTER_FILE)
master['Gene_Symbol_clean'] = master['Gene_Symbol'].astype(str).str.strip().str.upper()
resource_order = ['LCPathways', 'KEGG', 'Reactome', 'WikiPathways', 'MSigDB Hallmark']
gene_sets = {}
for resource in resource_order:
    genes = set(master.loc[master['Database'] == resource, 'Gene_Symbol_clean'].dropna())
    gene_sets[resource] = genes
rows = []
for A, B in itertools.combinations(resource_order, 2):
    genes_A = gene_sets[A]
    genes_B = gene_sets[B]
    shared = len(genes_A & genes_B)
    union = len(genes_A | genes_B)
    rows.append({'Resource_A': A, 'Resource_B': B, 'Genes_A': len(genes_A), 'Genes_B': len(genes_B), 'Shared_Genes': shared, 'Jaccard': shared / union, 'Coverage_A_to_B': shared / len(genes_A), 'Coverage_B_to_A': shared / len(genes_B)})
S2 = pd.DataFrame(rows)
for col in ['Jaccard', 'Coverage_A_to_B', 'Coverage_B_to_A']:
    S2[col] = S2[col].round(4)
print('\nS2:')
display(S2)
S3_file = RESULTS_DIR / 'BRCA_PRISM_gene_exact_intersection_counts.csv'
S3 = pd.read_csv(S3_file)
print('\nS3:', S3.shape)
print(S3.columns.tolist())
S4_file = RESULTS_DIR / 'BRCA_PRISM_step6C_final_reviewed_pathway_family_layer.csv'
S4 = pd.read_csv(S4_file)
print('\nS4:', S4.shape)
print(S4.columns.tolist())
S5_file = RESULTS_DIR / 'BRCA_PRISM_step9B_PAM50_family_association.csv'
S5 = pd.read_csv(S5_file)
print('\nS5:', S5.shape)
print(S5.columns.tolist())
S7A_file = RESULTS_DIR / 'BRCA_PRISM_step16C_repeatedCV_ablation_summary.csv'
S7A = pd.read_csv(S7A_file)
print('\nS7A:', S7A.shape)
print(S7A.columns.tolist())
S7B_file = RESULTS_DIR / 'BRCA_PRISM_step16D_paired_ablation_statistics.csv'
S7B = pd.read_csv(S7B_file)
print('\nS7B:', S7B.shape)
print(S7B.columns.tolist())
S8_file = RESULTS_DIR / 'BRCA_PRISM_CPTAC_step7_subtype_pattern_concordance.csv'
S8 = pd.read_csv(S8_file)
print('\nS8:', S8.shape)
print(S8.columns.tolist())
S9_file = RESULTS_DIR / 'BRCA_PRISM_CPTAC_step9_external_ablation_summary.csv'
S9 = pd.read_csv(S9_file)
print('\nS9:', S9.shape)
print(S9.columns.tolist())
S1.to_csv(SUPP_DIR / 'Table_S1_manual_family_review.csv', index=False)
S2.to_csv(SUPP_DIR / 'Table_S2_pairwise_gene_overlap.csv', index=False)
S3.to_csv(SUPP_DIR / 'Table_S3_exact_gene_intersections.csv', index=False)
S4.to_csv(SUPP_DIR / 'Table_S4_pathway_family_catalog.csv', index=False)
S5.to_csv(SUPP_DIR / 'Table_S5_TCGA_PAM50_associations.csv', index=False)
S7A.to_csv(SUPP_DIR / 'Table_S7A_TCGA_ablation_performance.csv', index=False)
S7B.to_csv(SUPP_DIR / 'Table_S7B_TCGA_ablation_statistics.csv', index=False)
S8.to_csv(SUPP_DIR / 'Table_S8_CPTAC_subtype_pattern_concordance.csv', index=False)
S9.to_csv(SUPP_DIR / 'Table_S9_CPTAC_predictive_sensitivity.csv', index=False)
excel_file = SUPP_DIR / 'BRCA_PRISM_Supplementary_Tables.xlsx'
with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
    S1.to_excel(writer, sheet_name='Table_S1', index=False)
    S2.to_excel(writer, sheet_name='Table_S2', index=False)
    S3.to_excel(writer, sheet_name='Table_S3', index=False)
    S4.to_excel(writer, sheet_name='Table_S4', index=False)
    S5.to_excel(writer, sheet_name='Table_S5', index=False)
    S7A.to_excel(writer, sheet_name='Table_S7A', index=False)
    S7B.to_excel(writer, sheet_name='Table_S7B', index=False)
    S8.to_excel(writer, sheet_name='Table_S8', index=False)
    S9.to_excel(writer, sheet_name='Table_S9', index=False)
print('\n' + '=' * 80)
print('SUPPLEMENTARY FILES CREATED')
print('=' * 80)
print('\nFolder:')
print(SUPP_DIR)
print('\nExcel workbook:')
print(excel_file)
from pathlib import Path
import pandas as pd
import itertools
PROJECT_ROOT = Path(__file__).resolve().parents[1]
RESULTS_DIR = PROJECT_ROOT / 'results'
DATA_DIR = PROJECT_ROOT / 'data_processed'
SUPP_DIR = PROJECT_ROOT / 'supplementary'
SUPP_DIR.mkdir(parents=True, exist_ok=True)
MASTER_FILE = DATA_DIR / 'BRCA_PRISM_all_pathway_resources_master.csv'
print('=' * 90)
print('PROJECT ROOT:')
print(PROJECT_ROOT)
print('\nMASTER FILE:')
print(MASTER_FILE)
print('\nMASTER EXISTS:', MASTER_FILE.exists())
print('=' * 90)
S1_file = RESULTS_DIR / 'BRCA_PRISM_step6C_BRCA_family_manual_review_COMPLETED-(S1).csv'
assert S1_file.exists(), f'S1 file not found:\n{S1_file}'
S1 = pd.read_csv(S1_file)
print('\n' + '=' * 90)
print('TABLE S1')
print('=' * 90)
print('File:', S1_file.name)
print('Shape:', S1.shape)
print('Columns:')
print(S1.columns.tolist())
display(S1.head())
assert MASTER_FILE.exists(), f'Master file not found:\n{MASTER_FILE}'
master = pd.read_csv(MASTER_FILE)
print('\n' + '=' * 90)
print('MASTER DATASET')
print('=' * 90)
print('Original shape:', master.shape)
print('Columns:')
print(master.columns.tolist())
required_cols = ['Database', 'Gene_Symbol']
for col in required_cols:
    assert col in master.columns, f'Missing required column: {col}'
master = master.dropna(subset=['Database', 'Gene_Symbol']).copy()
master['Database'] = master['Database'].astype(str).str.strip()
master['Gene_Symbol_clean'] = master['Gene_Symbol'].astype(str).str.strip().str.upper()
master = master[(master['Gene_Symbol_clean'] != '') & (master['Gene_Symbol_clean'] != 'NAN')].copy()
print('Cleaned shape:', master.shape)
print('\nDatabase labels found:')
print(sorted(master['Database'].unique()))

def find_database_label(options):
    existing = set(master['Database'].unique())
    for option in options:
        if option in existing:
            return option
    raise ValueError(f'Could not find any of these database labels: {options}')
LC_LABEL = find_database_label(['LCPathways'])
KEGG_LABEL = find_database_label(['KEGG'])
REACTOME_LABEL = find_database_label(['Reactome'])
WIKI_LABEL = find_database_label(['WikiPathways'])
HALLMARK_LABEL = find_database_label(['MSigDB_Hallmark', 'MSigDB Hallmark'])
resource_order = [LC_LABEL, KEGG_LABEL, REACTOME_LABEL, WIKI_LABEL, HALLMARK_LABEL]
print('\nResolved resource order:')
for resource in resource_order:
    print(resource)
exact_gene_counts = master.groupby('Database')['Gene_Symbol_clean'].nunique().reindex(resource_order)
print('\n' + '=' * 90)
print('EXACT UNIQUE GENE COUNTS')
print('=' * 90)
print(exact_gene_counts)
print('\nReactome exact unique genes:', exact_gene_counts.loc[REACTOME_LABEL])
expected_counts = {LC_LABEL: 3162, KEGG_LABEL: 9262, REACTOME_LABEL: 11983, WIKI_LABEL: 7985, HALLMARK_LABEL: 4310}
print('\nValidation against manuscript values:')
for db, expected in expected_counts.items():
    observed = int(exact_gene_counts.loc[db])
    status = 'OK' if observed == expected else 'CHECK'
    print(f'{db:20s} observed={observed:6d} expected={expected:6d}  {status}')
gene_sets = {}
for resource in resource_order:
    genes = set(master.loc[master['Database'] == resource, 'Gene_Symbol_clean'].dropna().unique())
    gene_sets[resource] = genes
rows = []
for A, B in itertools.combinations(resource_order, 2):
    genes_A = gene_sets[A]
    genes_B = gene_sets[B]
    shared = len(genes_A.intersection(genes_B))
    union = len(genes_A.union(genes_B))
    jaccard = shared / union if union > 0 else 0
    coverage_A_to_B = shared / len(genes_A) if len(genes_A) > 0 else 0
    coverage_B_to_A = shared / len(genes_B) if len(genes_B) > 0 else 0
    rows.append({'Resource_A': A, 'Resource_B': B, 'Genes_A': len(genes_A), 'Genes_B': len(genes_B), 'Shared_Genes': shared, 'Jaccard': jaccard, 'Coverage_A_to_B': coverage_A_to_B, 'Coverage_B_to_A': coverage_B_to_A})
S2 = pd.DataFrame(rows)
numeric_cols = ['Jaccard', 'Coverage_A_to_B', 'Coverage_B_to_A']
for col in numeric_cols:
    S2[col] = S2[col].round(4)
print('\n' + '=' * 90)
print('TABLE S2')
print('=' * 90)
print('Shape:', S2.shape)
display(S2)
expected_jaccards = {frozenset([LC_LABEL, KEGG_LABEL]): 0.2338, frozenset([LC_LABEL, REACTOME_LABEL]): 0.2134, frozenset([LC_LABEL, WIKI_LABEL]): 0.2748, frozenset([LC_LABEL, HALLMARK_LABEL]): 0.2457, frozenset([KEGG_LABEL, REACTOME_LABEL]): 0.5249, frozenset([KEGG_LABEL, WIKI_LABEL]): 0.4822, frozenset([KEGG_LABEL, HALLMARK_LABEL]): 0.2913, frozenset([REACTOME_LABEL, WIKI_LABEL]): 0.4621, frozenset([REACTOME_LABEL, HALLMARK_LABEL]): 0.2851, frozenset([WIKI_LABEL, HALLMARK_LABEL]): 0.3292}
print('\nS2 Jaccard validation:')
for _, row in S2.iterrows():
    key = frozenset([row['Resource_A'], row['Resource_B']])
    expected = expected_jaccards.get(key)
    observed = float(row['Jaccard'])
    if expected is None:
        print(row['Resource_A'], row['Resource_B'], 'No expected value stored')
    else:
        status = 'OK' if abs(observed - expected) <= 0.0001 else 'CHECK'
        print(f"{row['Resource_A']} vs {row['Resource_B']}: observed={observed:.4f}, expected={expected:.4f} {status}")
S3_file = RESULTS_DIR / 'BRCA_PRISM_gene_exact_intersection_counts.csv'
assert S3_file.exists(), f'S3 file not found:\n{S3_file}'
S3 = pd.read_csv(S3_file)
print('\n' + '=' * 90)
print('TABLE S3')
print('=' * 90)
print('File:', S3_file.name)
print('Shape:', S3.shape)
print('Columns:')
print(S3.columns.tolist())
display(S3.head(20))
S4_file = RESULTS_DIR / 'BRCA_PRISM_step6C_final_reviewed_pathway_family_layer.csv'
assert S4_file.exists(), f'S4 file not found:\n{S4_file}'
S4 = pd.read_csv(S4_file)
print('\n' + '=' * 90)
print('TABLE S4')
print('=' * 90)
print('File:', S4_file.name)
print('Shape:', S4.shape)
print('Columns:')
print(S4.columns.tolist())
display(S4.head(20))
S5_file = RESULTS_DIR / 'BRCA_PRISM_step9B_PAM50_family_association.csv'
assert S5_file.exists(), f'S5 file not found:\n{S5_file}'
S5 = pd.read_csv(S5_file)
print('\n' + '=' * 90)
print('TABLE S5')
print('=' * 90)
print('File:', S5_file.name)
print('Shape:', S5.shape)
print('Columns:')
print(S5.columns.tolist())
display(S5.head(31))
S7A_file = RESULTS_DIR / 'BRCA_PRISM_step16C_repeatedCV_ablation_summary.csv'
assert S7A_file.exists(), f'S7A file not found:\n{S7A_file}'
S7A = pd.read_csv(S7A_file)
print('\n' + '=' * 90)
print('TABLE S7A')
print('=' * 90)
print('File:', S7A_file.name)
print('Shape:', S7A.shape)
print('Columns:')
print(S7A.columns.tolist())
display(S7A)
S7B_file = RESULTS_DIR / 'BRCA_PRISM_step16D_paired_ablation_statistics.csv'
assert S7B_file.exists(), f'S7B file not found:\n{S7B_file}'
S7B = pd.read_csv(S7B_file)
print('\n' + '=' * 90)
print('TABLE S7B')
print('=' * 90)
print('File:', S7B_file.name)
print('Shape:', S7B.shape)
print('Columns:')
print(S7B.columns.tolist())
display(S7B)
S8_file = RESULTS_DIR / 'BRCA_PRISM_CPTAC_step7_subtype_pattern_concordance.csv'
assert S8_file.exists(), f'S8 file not found:\n{S8_file}'
S8 = pd.read_csv(S8_file)
print('\n' + '=' * 90)
print('TABLE S8')
print('=' * 90)
print('File:', S8_file.name)
print('Shape:', S8.shape)
print('Columns:')
print(S8.columns.tolist())
display(S8.head(31))
S9_file = RESULTS_DIR / 'BRCA_PRISM_CPTAC_step9_external_ablation_summary.csv'
assert S9_file.exists(), f'S9 file not found:\n{S9_file}'
S9 = pd.read_csv(S9_file)
print('\n' + '=' * 90)
print('TABLE S9')
print('=' * 90)
print('File:', S9_file.name)
print('Shape:', S9.shape)
print('Columns:')
print(S9.columns.tolist())
display(S9)
print('\n' + '=' * 90)
print('SAVING SUPPLEMENTARY CSV FILES')
print('=' * 90)
S1_out = SUPP_DIR / 'Table_S1_manual_family_expansion_review.csv'
S2_out = SUPP_DIR / 'Table_S2_pairwise_gene_overlap.csv'
S3_out = SUPP_DIR / 'Table_S3_exact_gene_intersections.csv'
S4_out = SUPP_DIR / 'Table_S4_complete_pathway_family_catalog.csv'
S5_out = SUPP_DIR / 'Table_S5_TCGA_PAM50_family_associations.csv'
S7A_out = SUPP_DIR / 'Table_S7A_TCGA_ablation_performance.csv'
S7B_out = SUPP_DIR / 'Table_S7B_TCGA_ablation_statistics.csv'
S8_out = SUPP_DIR / 'Table_S8_CPTAC_subtype_pattern_concordance.csv'
S9_out = SUPP_DIR / 'Table_S9_CPTAC_predictive_sensitivity.csv'
S1.to_csv(S1_out, index=False)
S2.to_csv(S2_out, index=False)
S3.to_csv(S3_out, index=False)
S4.to_csv(S4_out, index=False)
S5.to_csv(S5_out, index=False)
S7A.to_csv(S7A_out, index=False)
S7B.to_csv(S7B_out, index=False)
S8.to_csv(S8_out, index=False)
S9.to_csv(S9_out, index=False)
excel_file = SUPP_DIR / 'BRCA_PRISM_Supplementary_Tables_S1-S5_S7-S9.xlsx'
with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
    S1.to_excel(writer, sheet_name='Table_S1', index=False)
    S2.to_excel(writer, sheet_name='Table_S2', index=False)
    S3.to_excel(writer, sheet_name='Table_S3', index=False)
    S4.to_excel(writer, sheet_name='Table_S4', index=False)
    S5.to_excel(writer, sheet_name='Table_S5', index=False)
    S7A.to_excel(writer, sheet_name='Table_S7A', index=False)
    S7B.to_excel(writer, sheet_name='Table_S7B', index=False)
    S8.to_excel(writer, sheet_name='Table_S8', index=False)
    S9.to_excel(writer, sheet_name='Table_S9', index=False)
print('\n' + '=' * 90)
print('SUPPLEMENTARY TABLE BUILD COMPLETE')
print('=' * 90)
print('\nSupplementary folder:')
print(SUPP_DIR)
print('\nCreated files:')
print(S1_out.name)
print(S2_out.name)
print(S3_out.name)
print(S4_out.name)
print(S5_out.name)
print(S7A_out.name)
print(S7B_out.name)
print(S8_out.name)
print(S9_out.name)
print('\nExcel workbook:')
print(excel_file)
print('\nNOTE:')
print('Table S6 has NOT yet been added.')
print('Next step: inspect BRCA_PRISM_step14C_gene_importance_rankings.csv.')
import pandas as pd
S4_file = RESULTS_DIR / 'BRCA_PRISM_step6C_final_reviewed_pathway_family_layer.csv'
family_summary_file = RESULTS_DIR / 'BRCA_PRISM_step6C_final_reviewed_family_summary.csv'
S4_raw = pd.read_csv(S4_file)
family_summary = pd.read_csv(family_summary_file)
print('S4 columns:')
print(S4_raw.columns.tolist())
print('\nFamily summary columns:')
print(family_summary.columns.tolist())
display(family_summary.head())
from pathlib import Path
import pandas as pd
print('=' * 90)
print('SEARCHING FOR FILES WITH NON-EMPTY Family_Name')
print('=' * 90)
matches = []
for f in RESULTS_DIR.glob('*.csv'):
    try:
        df = pd.read_csv(f)
        if 'Family_Name' in df.columns:
            nonempty = df['Family_Name'].notna().sum()
            if nonempty > 0:
                matches.append((f.name, df.shape, nonempty))
                print('\nFILE:', f.name)
                print('Shape:', df.shape)
                print('Non-empty Family_Name:', nonempty)
                display(df[['Family_ID', 'Family_Name']].dropna().drop_duplicates().head(15))
    except Exception:
        pass
print('\n' + '=' * 90)
print('SUMMARY')
print('=' * 90)
for item in matches:
    print(item)
import pandas as pd
print('=' * 100)
print('SEARCHING ALL RESULTS FILES FOR FAMILY DESCRIPTION COLUMNS')
print('=' * 100)
keywords = ['name', 'label', 'representative', 'pathway', 'theme', 'description']
for f in sorted(RESULTS_DIR.glob('*.csv')):
    try:
        df = pd.read_csv(f)
        if 'Family_ID' not in df.columns:
            continue
        candidate_cols = [c for c in df.columns if any((k in c.lower() for k in keywords))]
        if candidate_cols:
            print('\nFILE:', f.name)
            print('Shape:', df.shape)
            print('Possible descriptive columns:')
            print(candidate_cols)
            cols_to_show = ['Family_ID'] + candidate_cols
            display(df[cols_to_show].drop_duplicates().head(15))
    except Exception as e:
        pass
S6_source = RESULTS_DIR / 'BRCA_PRISM_step14C_gene_importance_rankings.csv'
gene_imp = pd.read_csv(S6_source)
print('=' * 100)
print('S6 SOURCE')
print('=' * 100)
print('Shape:', gene_imp.shape)
print('\nColumns:')
print(gene_imp.columns.tolist())
print('\nFirst 20 rows:')
display(gene_imp.head(20))
from pathlib import Path
import pandas as pd
PROJECT_ROOT = Path(__file__).resolve().parents[1]
RESULTS_DIR = PROJECT_ROOT / 'results'
SUPP_DIR = PROJECT_ROOT / 'supplementary'
SUPP_DIR.mkdir(parents=True, exist_ok=True)
S4_layer_file = RESULTS_DIR / 'BRCA_PRISM_step6C_final_reviewed_pathway_family_layer.csv'
family_summary_file = RESULTS_DIR / 'BRCA_PRISM_step6C_final_reviewed_family_summary.csv'
assert S4_layer_file.exists(), f'File not found:\n{S4_layer_file}'
assert family_summary_file.exists(), f'File not found:\n{family_summary_file}'
layer = pd.read_csv(S4_layer_file)
family_summary = pd.read_csv(family_summary_file)
print('=' * 90)
print('SOURCE FILES')
print('=' * 90)
print('\nPathway-family layer shape:', layer.shape)
print('Columns:')
print(layer.columns.tolist())
print('\nFamily summary shape:', family_summary.shape)
print('Columns:')
print(family_summary.columns.tolist())
S4 = layer.copy()
if 'Family_Name' in S4.columns:
    nonmissing = S4['Family_Name'].notna().sum()
    print('\nNon-empty Family_Name values:', nonmissing)
    if nonmissing == 0:
        S4 = S4.drop(columns=['Family_Name'])
        print('Blank Family_Name column removed.')
summary_cols = ['Family_ID', 'Number_of_Pathways', 'Number_of_Resources', 'Resources']
missing_cols = [c for c in summary_cols if c not in family_summary.columns]
if missing_cols:
    raise ValueError(f'Missing columns in family summary: {missing_cols}')
S4 = S4.merge(family_summary[summary_cols], on='Family_ID', how='left', validate='many_to_one')
final_columns = ['Family_ID', 'Number_of_Pathways', 'Number_of_Resources', 'Resources', 'Database', 'Pathway_ID', 'Pathway_Name', 'Gene_Count', 'Family_Membership_Evidence', 'Original_Pathway_Preserved']
S4 = S4[final_columns]
S4 = S4.sort_values(['Family_ID', 'Database', 'Pathway_ID'], na_position='last').reset_index(drop=True)
print('\n' + '=' * 90)
print('S4 VALIDATION')
print('=' * 90)
print('Total pathway records:', len(S4))
print('Unique families:', S4['Family_ID'].nunique())
print('Missing Family_ID:', S4['Family_ID'].isna().sum())
print('Missing Database:', S4['Database'].isna().sum())
print('Missing Pathway_ID:', S4['Pathway_ID'].isna().sum())
print('Missing Pathway_Name:', S4['Pathway_Name'].isna().sum())
print('\nOriginal_Pathway_Preserved:')
print(S4['Original_Pathway_Preserved'].value_counts(dropna=False))
assert len(S4) == 116, f'Expected 116 pathway records, found {len(S4)}'
assert S4['Family_ID'].nunique() == 31, f"Expected 31 families, found {S4['Family_ID'].nunique()}"
assert S4['Number_of_Pathways'].notna().all()
assert S4['Number_of_Resources'].notna().all()
assert S4['Resources'].notna().all()
print('\nValidation PASSED.')
print('\n' + '=' * 90)
print('CLEAN TABLE S4 PREVIEW')
print('=' * 90)
display(S4.head(25))
S4_csv = SUPP_DIR / 'Table_S4_complete_pathway_family_catalog.csv'
S4.to_csv(S4_csv, index=False)
S4_excel = SUPP_DIR / 'Table_S4_complete_pathway_family_catalog.xlsx'
S4.to_excel(S4_excel, index=False, sheet_name='Table_S4')
print('\n' + '=' * 90)
print('TABLE S4 COMPLETE')
print('=' * 90)
print('\nCSV:')
print(S4_csv)
print('\nExcel:')
print(S4_excel)
import pandas as pd
import numpy as np
import itertools
from scipy.stats import spearmanr
S6_source = RESULTS_DIR / 'BRCA_PRISM_step14C_gene_importance_rankings.csv'
gene_imp = pd.read_csv(S6_source)
print('Source shape:', gene_imp.shape)
print('Columns:', gene_imp.columns.tolist())
gene_imp = gene_imp[['Representation', 'Gene_Symbol', 'Gene_Importance', 'Number_of_Important_Families', 'Gene_Rank']].copy()
gene_imp = gene_imp.dropna(subset=['Representation', 'Gene_Symbol', 'Gene_Importance'])
gene_imp['Gene_Symbol'] = gene_imp['Gene_Symbol'].astype(str).str.strip().str.upper()
gene_imp['Gene_Importance'] = pd.to_numeric(gene_imp['Gene_Importance'], errors='coerce')
gene_imp['Gene_Rank'] = pd.to_numeric(gene_imp['Gene_Rank'], errors='coerce')
gene_imp = gene_imp.dropna(subset=['Gene_Importance'])
representations = gene_imp['Representation'].dropna().unique().tolist()
print('\nRepresentations found:')
for r in representations:
    print(' -', r)
print('\nNumber of representations:', len(representations))
rows = []
for rep_A, rep_B in itertools.combinations(representations, 2):
    A = gene_imp[gene_imp['Representation'] == rep_A][['Gene_Symbol', 'Gene_Importance', 'Gene_Rank']].copy()
    B = gene_imp[gene_imp['Representation'] == rep_B][['Gene_Symbol', 'Gene_Importance', 'Gene_Rank']].copy()
    A = A.rename(columns={'Gene_Importance': 'Gene_Importance_A', 'Gene_Rank': 'Gene_Rank_A'})
    B = B.rename(columns={'Gene_Importance': 'Gene_Importance_B', 'Gene_Rank': 'Gene_Rank_B'})
    merged = A.merge(B, on='Gene_Symbol', how='inner')
    shared_genes = len(merged)
    if shared_genes >= 2:
        rho, p_value = spearmanr(merged['Gene_Importance_A'], merged['Gene_Importance_B'])
    else:
        rho = np.nan
        p_value = np.nan
    top20_A = set(A.sort_values(['Gene_Rank_A', 'Gene_Symbol']).head(20)['Gene_Symbol'])
    top20_B = set(B.sort_values(['Gene_Rank_B', 'Gene_Symbol']).head(20)['Gene_Symbol'])
    top20_intersection = top20_A & top20_B
    top20_union = top20_A | top20_B
    top20_shared = len(top20_intersection)
    top20_jaccard = top20_shared / len(top20_union) if len(top20_union) > 0 else np.nan
    rows.append({'Representation_A': rep_A, 'Representation_B': rep_B, 'Genes_in_A': len(A), 'Genes_in_B': len(B), 'Shared_Genes': shared_genes, 'Spearman_Rho': rho, 'Spearman_P_Value': p_value, 'Top20_Shared_Genes': top20_shared, 'Top20_Jaccard': top20_jaccard})
S6 = pd.DataFrame(rows)
S6['Spearman_Rho'] = S6['Spearman_Rho'].round(4)
S6['Top20_Jaccard'] = S6['Top20_Jaccard'].round(4)
print('\n' + '=' * 100)
print('TABLE S6')
print('=' * 100)
display(S6)
S6_out = SUPP_DIR / 'Table_S6_gene_importance_pairwise_comparisons.csv'
S6.to_csv(S6_out, index=False)
print('\nSaved:')
print(S6_out)
print('\n' + '=' * 100)
print('S6 VALIDATION')
print('=' * 100)
print('Number of pairwise comparisons:', len(S6))
print('\nSpearman values:')
print(S6[['Representation_A', 'Representation_B', 'Spearman_Rho']])
print('\nTop-20 Jaccard values:')
print(S6[['Representation_A', 'Representation_B', 'Top20_Jaccard']])
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
PROJECT_ROOT = Path(__file__).resolve().parents[1]
SUPP_DIR = PROJECT_ROOT / 'supplementary'
files = {'Table_S1': SUPP_DIR / 'Table_S1_manual_family_expansion_review.csv', 'Table_S2': SUPP_DIR / 'Table_S2_pairwise_gene_overlap.csv', 'Table_S3': SUPP_DIR / 'Table_S3_exact_gene_intersections.csv', 'Table_S4': SUPP_DIR / 'Table_S4_complete_pathway_family_catalog.csv', 'Table_S5': SUPP_DIR / 'Table_S5_TCGA_PAM50_family_associations.csv', 'Table_S6': SUPP_DIR / 'Table_S6_gene_importance_pairwise_comparisons.csv', 'Table_S7A': SUPP_DIR / 'Table_S7A_TCGA_ablation_performance.csv', 'Table_S7B': SUPP_DIR / 'Table_S7B_TCGA_ablation_statistics.csv', 'Table_S8': SUPP_DIR / 'Table_S8_CPTAC_subtype_pattern_concordance.csv', 'Table_S9': SUPP_DIR / 'Table_S9_CPTAC_predictive_sensitivity.csv'}
print('=' * 100)
print('CHECKING FINAL SUPPLEMENTARY FILES')
print('=' * 100)
for sheet, file_path in files.items():
    status = 'FOUND' if file_path.exists() else 'MISSING'
    print(f'{sheet:10s}  {status:8s}  {file_path.name}')
missing = [str(path) for path in files.values() if not path.exists()]
if missing:
    raise FileNotFoundError('\nMissing supplementary files:\n' + '\n'.join(missing))
tables = {}
for sheet, file_path in files.items():
    tables[sheet] = pd.read_csv(file_path)
    print(f'{sheet}: {tables[sheet].shape[0]} rows x {tables[sheet].shape[1]} columns')
FINAL_EXCEL = SUPP_DIR / 'BRCA_PRISM_Supplementary_Tables.xlsx'
with pd.ExcelWriter(FINAL_EXCEL, engine='openpyxl') as writer:
    for sheet_name, df in tables.items():
        df.to_excel(writer, sheet_name=sheet_name, index=False)
wb = load_workbook(FINAL_EXCEL)
for ws in wb.worksheets:
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in column_cells:
            value = '' if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        width = min(max(max_length + 2, 10), 45)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30
wb.save(FINAL_EXCEL)
print('\n' + '=' * 100)
print('FINAL WORKBOOK CREATED')
print('=' * 100)
print('\nFile:')
print(FINAL_EXCEL)
print('\nSheets:')
for ws in wb.worksheets:
    print(f'{ws.title:10s} : {ws.max_row - 1} data rows, {ws.max_column} columns')
print('\nDONE.')
S7C_file = RESULTS_DIR / 'BRCA_PRISM_step16B_ablation_gene_counts.csv'
S7C = pd.read_csv(S7C_file)
print('S7C shape:', S7C.shape)
print('Columns:')
print(S7C.columns.tolist())
display(S7C)
